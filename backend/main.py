# --- Import section ---
# Core FastAPI components
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, status, Body
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import delete
from backend.database import AsyncSessionLocal, engine
from backend import models
from pydantic import BaseModel
import re
from typing import List, Dict, Any, Optional
import csv
from io import StringIO, BytesIO
from openpyxl import load_workbook, Workbook
from fastapi.responses import JSONResponse, FileResponse
import os
import tempfile

# --- Database initialization ---
# This function runs at app startup to create database tables
async def init_models():
    async with engine.begin() as conn:
        await conn.run_sync(models.Base.metadata.create_all)

# Create our FastAPI application
app = FastAPI(on_startup=[init_models])

# --- CORS configuration ---
# Allow frontend requests from localhost
origins = [
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Database session helper ---
# Creates a new database connection for each request
async def get_db():
    async with AsyncSessionLocal() as session:
        yield session

# --- Base health check endpoint ---
@app.get("/")
async def read_root():
    return {"message": "Welcome to the FastAPI Backend!"}

# --- Asset Management Endpoints ---

# Get all assets in the system
@app.get("/api/assets")
async def get_assets(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Assets))
    assets = result.scalars().all()
    return assets

# Get a single asset by ID
@app.get("/api/assets/{asset_id}")
async def get_asset(asset_id: int, db: AsyncSession = Depends(get_db)):
    print(f"Fetching asset with asset_id: {asset_id}")
    result = await db.execute(select(models.Assets).where(models.Assets.asset_id == asset_id))
    asset = result.scalars().first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return {
        "asset_id": asset.asset_id,
        "asset_name": asset.asset_name,
        "asset_type": asset.asset_type,
    }

# Get all subgroups that belong to a specific asset
@app.get("/api/assets/{asset_id}/subgroups")
async def get_subgroups(asset_id: int, db: AsyncSession = Depends(get_db)):
    # First check if the asset exists
    asset_result = await db.execute(select(models.Assets).where(models.Assets.asset_id == asset_id))
    asset = asset_result.scalars().first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    # Get subgroups (may be empty)
    result = await db.execute(select(models.Subgroups).where(models.Subgroups.asset_id == asset_id))
    subgroups = result.scalars().all()
    return subgroups

# Get details for a specific subgroup
@app.get("/api/subgroups/{subgroup_id}")
async def get_subgroup(subgroup_id: int, db: AsyncSession = Depends(get_db)):
    print(f"Fetching subgroup for subgroup_id: {subgroup_id}")
    result = await db.execute(select(models.Subgroups).where(models.Subgroups.subgroup_id == subgroup_id))
    subgroup = result.scalars().first()
    if not subgroup:
        print(f"No subgroup found for subgroup_id: {subgroup_id}")
        raise HTTPException(status_code=404, detail="Subgroup not found")
    print(f"Found subgroup: {subgroup}")
    return subgroup

# Data validation model for creating assets
class AssetCreate(BaseModel):
    asset_name: str
    asset_type: str

# Create a new asset
@app.post("/api/assets")
async def create_asset(asset: AssetCreate, db: AsyncSession = Depends(get_db)):
    new_asset = models.Assets(asset_name=asset.asset_name, asset_type=asset.asset_type)
    db.add(new_asset)
    await db.commit()
    await db.refresh(new_asset)
    return new_asset

# Data validation for renaming an asset
class AssetRename(BaseModel):
    asset_name: str

# Rename an existing asset
@app.put("/api/assets/{asset_id}")
async def rename_asset(asset_id: int, asset: AssetRename, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Assets).where(models.Assets.asset_id == asset_id))
    existing_asset = result.scalars().first()
    if not existing_asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    existing_asset.asset_name = asset.asset_name
    await db.commit()
    await db.refresh(existing_asset)
    return existing_asset

# Data validation for creating subgroups
class SubgroupCreate(BaseModel):
    subgroup_name: str

# Create a new subgroup within an asset
@app.post("/api/assets/{asset_id}/subgroups")
async def create_subgroup(asset_id: int, subgroup: SubgroupCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Assets).where(models.Assets.asset_id == asset_id))
    existing_asset = result.scalars().first()
    if not existing_asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    new_subgroup = models.Subgroups(
        subgroup_name=subgroup.subgroup_name,
        asset_id=asset_id
    )
    db.add(new_subgroup)
    await db.commit()
    await db.refresh(new_subgroup)
    return new_subgroup

# Data validation for renaming subgroups
class SubgroupRename(BaseModel):
    subgroup_name: str

# Rename an existing subgroup
@app.put("/api/subgroups/{subgroup_id}")
async def rename_subgroup(subgroup_id: int, subgroup: SubgroupRename, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Subgroups).where(models.Subgroups.subgroup_id == subgroup_id))
    existing_subgroup = result.scalars().first()
    if not existing_subgroup:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    existing_subgroup.subgroup_name = subgroup.subgroup_name
    await db.commit()
    await db.refresh(existing_subgroup)
    return existing_subgroup

# Data validation for tags
class Tag(BaseModel):
    tag_name: str

# --- Masterlist Management ---

# Upload a masterlist file (CSV or Excel) and extract tags
@app.post("/api/upload_masterlist")
async def upload_masterlist(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    # Validate file type
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type")

    try:
        tags_column = "tags"
        content = await file.read()

        # Handle CSV files
        if file.filename.endswith('.csv'):
            content_str = content.decode("utf-8")
            reader = csv.DictReader(StringIO(content_str))

            async with db.begin():
                masterlist = models.MasterList(file_name=file.filename)
                db.add(masterlist)
                await db.flush()
                print(f"Added masterlist: {masterlist}")

                for row in reader:
                    print(f"Processing row: {row}")
                    if tags_column in row and row[tags_column] is not None:
                        # Split tags by comma and process each one
                        tags = row[tags_column].split(',')
                        for tag_name in tags:
                            tag = models.Tags(tag_name=tag_name.strip(), file_id=masterlist.file_id, tag_type="default")
                            db.add(tag)
                            print(f"Added tag: {tag}")
                    else:
                        print(f"Tags column '{tags_column}' not found or empty in the file")
                        raise HTTPException(status_code=400, detail=f"Tags column '{tags_column}' not found or empty in the file")

        # Handle Excel files
        else:
            workbook = load_workbook(filename=BytesIO(content))
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            
            # Find tags column index
            tags_idx = headers.index(tags_column) if tags_column in headers else None
            if tags_idx is None:
                print(f"Tags column '{tags_column}' not found in the file")
                raise HTTPException(status_code=400, detail=f"Tags column '{tags_column}' not found in the file")

            async with db.begin():
                masterlist = models.MasterList(file_name=file.filename)
                db.add(masterlist)
                await db.flush()
                print(f"Added masterlist: {masterlist}")

                # Process each row starting from row 2 (after headers)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    print(f"Processing row: {row}")
                    if row[tags_idx] is not None:
                        tags = row[tags_idx].split(',')
                        for tag_name in tags:
                            tag = models.Tags(tag_name=tag_name.strip(), file_id=masterlist.file_id, tag_type="default")
                            db.add(tag)
                            print(f"Added tag: {tag}")
                    else:
                        print(f"Tags column '{tags_column}' is empty in the file")
                        raise HTTPException(status_code=400, detail=f"Tags column '{tags_column}' is empty in the file")

        await db.commit()
        return {"message": "Masterlist uploaded successfully", "file_id": masterlist.file_id, "file_name": masterlist.file_name}

    except Exception as e:
        await db.rollback()
        print(f"Error occurred: {str(e)}")
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    
# Get all tags from a specific masterlist
@app.get("/api/tags")
async def get_tags_by_file_id(file_id: int, db: AsyncSession = Depends(get_db)):
    # First check if the file exists
    file_result = await db.execute(select(models.MasterList).where(models.MasterList.file_id == file_id))
    file = file_result.scalars().first()
    if not file:
        raise HTTPException(status_code=404, detail="Masterlist file not found")
    
    # Get tags (may be empty)
    result = await db.execute(select(models.Tags).where(models.Tags.file_id == file_id))
    tags = result.scalars().all()
    return tags

# Get the most recently uploaded masterlist
@app.get("/api/masterlist/latest")
async def get_latest_masterlist(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.MasterList).order_by(models.MasterList.file_id.desc()).limit(1))
    masterlist = result.scalars().first()
    if not masterlist:
        raise HTTPException(status_code=404, detail="No masterlist found")
    return {"file_id": masterlist.file_id, "file_name": masterlist.file_name}

# Get masterlist by ID
@app.get("/api/masterlist/{file_id}")
async def get_masterlist_by_file_id(file_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.MasterList).where(models.MasterList.file_id == file_id))
    masterlist = result.scalars().first()
    if not masterlist:
        raise HTTPException(status_code=404, detail="Masterlist not found for the given file ID")
    return {"file_id": masterlist.file_id, "file_name": masterlist.file_name}

# --- Subgroup Tag Management ---

# Data validation for creating subgroup tags
class SubgroupTagCreate(BaseModel):
    tag_id: int
    tag_name: str
    parent_subgroup_tag_id: int = None
    formula_id: int = None  # Optional formula ID

# Add a tag to a subgroup or as a child of another tag
@app.post("/api/subgroups/{subgroup_id}/tags", status_code=status.HTTP_201_CREATED)
async def add_tag_to_subgroup(subgroup_id: int, tag: SubgroupTagCreate, db: AsyncSession = Depends(get_db)):
    print(f"Received request to add tag {tag.tag_id} to subgroup {subgroup_id}")
    try:
        # Check if formula exists if formula_id is provided
        if tag.formula_id:
            formula_result = await db.execute(select(models.Formulas).where(models.Formulas.formula_id == tag.formula_id))
            formula = formula_result.scalars().first()
            if not formula:
                return JSONResponse(status_code=404, content={"detail": "Formula not found"})
                
        # Only verify the subgroup exists if this is a root-level tag (no parent)
        if tag.parent_subgroup_tag_id is None:
            result = await db.execute(select(models.Subgroups).where(models.Subgroups.subgroup_id == subgroup_id))
            existing_subgroup = result.scalars().first()
            if not existing_subgroup:
                return JSONResponse(status_code=404, content={"detail": "Subgroup not found"})
            
            # This is a root-level tag, associate it with the subgroup
            new_subgroup_tag = models.SubgroupTag(
                subgroup_id=subgroup_id,
                tag_id=tag.tag_id,
                subgroup_tag_name=tag.tag_name,
                parent_subgroup_tag_id=None,
                formula_id=tag.formula_id
            )
        else:
            # This is a child tag, only associate it with the parent tag
            # First, verify the parent tag exists
            result = await db.execute(
                select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == tag.parent_subgroup_tag_id)
            )
            parent_tag = result.scalars().first()
            if not parent_tag:
                return JSONResponse(status_code=404, content={"detail": "Parent tag not found"})
            
            # Create child tag without subgroup_id
            new_subgroup_tag = models.SubgroupTag(
                subgroup_id=None,  # No direct association with subgroup
                tag_id=tag.tag_id,
                subgroup_tag_name=tag.tag_name,
                parent_subgroup_tag_id=tag.parent_subgroup_tag_id,
                formula_id=tag.formula_id
            )

        db.add(new_subgroup_tag)
        await db.commit()
        await db.refresh(new_subgroup_tag)
        print(f"Added tag {tag.tag_id} to {'subgroup ' + str(subgroup_id) if tag.parent_subgroup_tag_id is None else 'parent tag ' + str(tag.parent_subgroup_tag_id)}")
        return new_subgroup_tag
    except Exception as e:
        print(f"Error adding tag to subgroup: {e}")
        return JSONResponse(status_code=500, content={"detail": f"Internal Server Error: {str(e)}"})
    
# Get all tags for a specific subgroup
@app.get("/api/subgroups/{subgroup_id}/tags")
async def get_subgroup_tags(subgroup_id: int, db: AsyncSession = Depends(get_db)):
    # First check if the subgroup exists
    subgroup_result = await db.execute(select(models.Subgroups).where(models.Subgroups.subgroup_id == subgroup_id))
    subgroup = subgroup_result.scalars().first()
    if not subgroup:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    # Get tags but filter out template context tags
    result = await db.execute(
        select(models.SubgroupTag).where(
            models.SubgroupTag.subgroup_id == subgroup_id,
            ~models.SubgroupTag.subgroup_tag_name.like("__TEMPLATE_CONTEXT__%")
        )
    )
    tags = result.scalars().all()
    return tags

# Get all masterlist files
@app.get("/api/masterlists")
async def get_all_masterlists(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.MasterList))
    masterlists = result.scalars().all()
    return masterlists  # Might be an empty list if none exist

# --- Formula Management ---

# Data validation for creating formulas
class FormulaCreate(BaseModel):
    formula_name: str
    formula_desc: str = None
    formula_expression: str
    num_parameters: int

# Data validation for formula evaluation
class FormulaEvaluationRequest(BaseModel):
    formula_id: int
    parameters: Dict[str, Any]

# Get all formulas
@app.get("/api/formulas")
async def get_formulas(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Formulas))
    formulas = result.scalars().all()
    return formulas

# Get a specific formula by ID
@app.get("/api/formulas/{formula_id}")
async def get_formula(formula_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Formulas).where(models.Formulas.formula_id == formula_id))
    formula = result.scalars().first()
    if not formula:
        raise HTTPException(status_code=404, detail="Formula not found")
    return formula

# Data validation for creating formulas - remove num_parameters
class FormulaCreate(BaseModel):
    formula_name: str
    formula_desc: str = None
    formula_expression: str

# Helper function to extract variables from formula expression
def extract_variables(formula_expression):
    # Find all instances of $variable in the expression
    # The pattern finds any $ followed by word characters (letters, numbers, underscore)
    variables = re.findall(r'\$([a-zA-Z_][a-zA-Z0-9_]*)', formula_expression)
    # Return unique variable names without the $ sign
    return list(set(variables))

# Create a new formula
@app.post("/api/formulas")
async def create_formula(formula: FormulaCreate, db: AsyncSession = Depends(get_db)):
    async with db.begin():
        new_formula = models.Formulas(
            formula_name=formula.formula_name,
            formula_desc=formula.formula_desc,
            formula_expression=formula.formula_expression
        )
        db.add(new_formula)
        await db.flush()  # Get the formula_id
        
        # Rest of the function stays the same...  # Get the formula_id
        
        # Extract variables from the formula expression
        variable_names = extract_variables(formula.formula_expression)
        
        # Create variable records
        for var_name in variable_names:
            variable = models.FormulaVariable(
                formula_id=new_formula.formula_id,
                variable_name=var_name  # Store without the $ prefix
            )
            db.add(variable)
    
    await db.refresh(new_formula)
    
    # Fetch variables to include in response
    variables_stmt = select(models.FormulaVariable).where(models.FormulaVariable.formula_id == new_formula.formula_id)
    variables_result = await db.execute(variables_stmt)
    variables = variables_result.scalars().all()
    
    return {**new_formula.__dict__, "variables": [{"variable_id": v.variable_id, "variable_name": v.variable_name} for v in variables]}

# Update an existing formula
@app.put("/api/formulas/{formula_id}")
async def update_formula(formula_id: int, formula: FormulaCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Formulas).where(models.Formulas.formula_id == formula_id))
    existing_formula = result.scalars().first()
    if not existing_formula:
        raise HTTPException(status_code=404, detail="Formula not found")
    
    async with db.begin():
        # Update formula fields
        existing_formula.formula_name = formula.formula_name
        existing_formula.formula_desc = formula.formula_desc
        existing_formula.formula_expression = formula.formula_expression
        
        # Handle variables - first delete all existing
        await db.execute(
            delete(models.FormulaVariable)
            .where(models.FormulaVariable.formula_id == formula_id)
        )
        
        # Extract variables from the updated formula expression
        variable_names = extract_variables(formula.formula_expression)
        
        # Create new variable records
        for var_name in variable_names:
            new_var = models.FormulaVariable(
                formula_id=formula_id,
                variable_name=var_name  # Store without the $ prefix
            )
            db.add(new_var)
    
    # Fetch updated variables to include in response
    variables_stmt = select(models.FormulaVariable).where(models.FormulaVariable.formula_id == formula_id)
    variables_result = await db.execute(variables_stmt)
    variables = variables_result.scalars().all()
    
    return {**existing_formula.__dict__, "variables": [{"variable_id": v.variable_id, "variable_name": v.variable_name} for v in variables]}
# Delete a formula
@app.delete("/api/formulas/{formula_id}")
async def delete_formula(formula_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Formulas).where(models.Formulas.formula_id == formula_id))
    formula = result.scalars().first()
    if not formula:
        raise HTTPException(status_code=404, detail="Formula not found")
    
    await db.delete(formula)
    await db.commit()
    return {"message": "Formula deleted successfully"}

# Response model for variables with tag info
class VariableWithTagResponse(BaseModel):
    variable_id: int
    variable_name: str
    subgroup_tag_id: Optional[int] = None
    subgroup_tag_name: Optional[str] = None

    class Config:
        orm_mode = True

class VariableMappingRequest(BaseModel):
    subgroup_tag_id: int
# Map a variable to a tag
@app.put("/api/formula-variables/{variable_id}/map")
async def map_variable_to_tag(
    variable_id: int, 
    mapping: VariableMappingRequest,  # Use the Pydantic model instead of Body(...)
    db: AsyncSession = Depends(get_db)
):
    # Find the variable
    var_result = await db.execute(
        select(models.FormulaVariable).where(models.FormulaVariable.variable_id == variable_id)
    )
    variable = var_result.scalars().first()
    if not variable:
        raise HTTPException(status_code=404, detail="Variable not found")
    
    # Find the tag - access the ID from the model
    tag_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
    )
    tag = tag_result.scalars().first()
    if not tag:
        raise HTTPException(status_code=404, detail="Subgroup tag not found")
    
    # Update the variable with the tag ID
    variable.subgroup_tag_id = mapping.subgroup_tag_id
    await db.commit()
    await db.refresh(variable)
    
    return {
        "variable_id": variable.variable_id,
        "variable_name": variable.variable_name,
        "subgroup_tag_id": variable.subgroup_tag_id
    }

#TENTATIVE, FOR FUTURE USE IN CASE NEEDED
# Evaluate formula (updated to use variable mappings)
@app.post("/api/formulas/evaluate")
async def evaluate_formula(formula_id: int, db: AsyncSession = Depends(get_db)):
    # Find the formula
    formula_result = await db.execute(select(models.Formulas).where(models.Formulas.formula_id == formula_id))
    formula = formula_result.scalars().first()
    if not formula:
        raise HTTPException(status_code=404, detail="Formula not found")
    
    # Get the expression
    expression = formula.formula_expression
    
    try:
        # Get all variables for this formula
        variables_result = await db.execute(
            select(models.FormulaVariable)
            .where(models.FormulaVariable.formula_id == formula.formula_id)
        )
        variables = variables_result.scalars().all()
        
        # Prepare evaluation environment
        eval_env = {}
        missing_tags = []
        
        for variable in variables:
            if not variable.subgroup_tag_id:
                missing_tags.append(variable.variable_name)
                continue
                
            # Get the tag value
            tag_result = await db.execute(
                select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == variable.subgroup_tag_id)
            )
            tag = tag_result.scalars().first()
            
            if not tag or tag.tag_value is None:
                missing_tags.append(variable.variable_name)
                continue
            
            # Add to evaluation environment
            var_name = variable.variable_name
            var_value = tag.tag_value
            
            # Try to convert value to appropriate type
            try:
                if isinstance(var_value, str):
                    if var_value.lower() == "true":
                        eval_env[var_name] = True
                    elif var_value.lower() == "false":
                        eval_env[var_name] = False
                    else:
                        # Try to convert to number
                        try:
                            eval_env[var_name] = float(var_value)
                        except:
                            eval_env[var_name] = var_value
                else:
                    eval_env[var_name] = var_value
            except:
                eval_env[var_name] = var_value
        
        # Check if we're missing any tag values
        if missing_tags:
            return {
                "formula_id": formula.formula_id,
                "error": f"Missing tag values for variables: {', '.join(missing_tags)}"
            }
        
        # Replace $variable with variable names in expression
        for var_name in eval_env:
            expression = expression.replace(f"${var_name}", var_name)
        
        # Evaluate the formula
        result = eval(expression, {"__builtins__": {}}, eval_env)
        return {
            "formula_id": formula.formula_id,
            "result": result
        }
    except Exception as e:
        return {
            "formula_id": formula.formula_id,
            "error": str(e)
        }
# --- Response Model for Subgroup Tags ---
class SubgroupTagResponse(BaseModel):
    subgroup_tag_id: int
    subgroup_tag_name: str
    parent_subgroup_tag_id: Optional[int]
    formula_id: Optional[int]

    class Config:
        orm_mode = True
        
# Get all child tags for a parent tag
@app.get("/api/subgroups/{subgroup_tag_id}/children_tags", response_model=List[SubgroupTagResponse])
async def get_children_tags(subgroup_tag_id: int, db: AsyncSession = Depends(get_db)):
    # First check if the parent tag exists
    parent_result = await db.execute(select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == subgroup_tag_id))
    parent = parent_result.scalars().first()
    if not parent:
        raise HTTPException(status_code=404, detail="Parent tag not found")
    
    # Get children tags (may be empty)
    result = await db.execute(select(models.SubgroupTag).where(models.SubgroupTag.parent_subgroup_tag_id == subgroup_tag_id))
    children_tags = result.scalars().all()
    return children_tags

# Update the formula attached to a subgroup tag
@app.put("/api/subgroups/{subgroup_tag_id}/formula")
async def update_subgroup_tag_formula(
    subgroup_tag_id: int,
    formula_data: dict,
    db: AsyncSession = Depends(get_db)
):
    # Update the formula for a subgroup tag
    async with db.begin():
        # Get the subgroup tag
        stmt = select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == subgroup_tag_id)
        result = await db.execute(stmt)
        subgroup_tag = result.scalar_one_or_none()
        
        if not subgroup_tag:
            raise HTTPException(status_code=404, detail=f"Subgroup tag with id {subgroup_tag_id} not found")
        
        # Update the formula ID
        subgroup_tag.formula_id = formula_data.get("formula_id")
        await db.commit()
    
    return {"message": "Formula assigned successfully"}

# Export subgroup tag data to Excel
@app.post("/api/subgroups/{subgroup_tag_id}/export")
async def export_subgroup_tag_data(subgroup_tag_id: int, db: AsyncSession = Depends(get_db)):
    # Get the parent tag
    stmt = select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == subgroup_tag_id)
    result = await db.execute(stmt)
    parent_tag = result.scalar_one_or_none()
    
    if not parent_tag:
        raise HTTPException(status_code=404, detail=f"Subgroup tag with id {subgroup_tag_id} not found")
    
    # Get formula if it exists
    formula = None
    if parent_tag.formula_id:
        formula_stmt = select(models.Formulas).where(models.Formulas.formula_id == parent_tag.formula_id)
        formula_result = await db.execute(formula_stmt)
        formula = formula_result.scalar_one_or_none()
    
    # Get child tags
    child_tags_stmt = select(models.SubgroupTag).where(models.SubgroupTag.parent_subgroup_tag_id == subgroup_tag_id)
    child_tags_result = await db.execute(child_tags_stmt)
    child_tags = child_tags_result.scalars().all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Subgroup Tag Data"
    
    # Add column headers
    headers = ["Parent Tag Name", "Formula Expression", "Child Tags"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add parent tag data
    ws.cell(row=2, column=1, value=parent_tag.subgroup_tag_name)
    ws.cell(row=2, column=2, value=formula.formula_expression if formula else "None")
    
    # Add child tags data
    if child_tags:
        for i, tag in enumerate(child_tags):
            ws.cell(row=i+2, column=3, value=tag.subgroup_tag_name)
    else:
        ws.cell(row=2, column=3, value="None")
    
    # Save to temporary file and return as download
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        file_path = tmp.name
    
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"subgroup_tag_{subgroup_tag_id}_export.xlsx"
    )

# Get all variables for a formula with their tag mappings
@app.get("/api/formulas/{formula_id}/variables", response_model=List[VariableWithTagResponse])
async def get_formula_variables(
    formula_id: int, 
    context_tag_id: Optional[int] = None,  # Add optional context parameter
    db: AsyncSession = Depends(get_db)
):
    # First check if the formula exists
    formula_result = await db.execute(select(models.Formulas).where(models.Formulas.formula_id == formula_id))
    formula = formula_result.scalars().first()
    if not formula:
        raise HTTPException(status_code=404, detail="Formula not found")
    
    # Get variables for this formula
    result = await db.execute(select(models.FormulaVariable).where(models.FormulaVariable.formula_id == formula_id))
    variables = result.scalars().all()
    
    # Get tag names for variables that have mappings
    response_variables = []
    for variable in variables:
        var_data = {
            "variable_id": variable.variable_id,
            "variable_name": variable.variable_name,
            "subgroup_tag_id": None,
            "subgroup_tag_name": None
        }
        
        # If context is provided, look up mappings from VariableTagMapping
        if context_tag_id:
            mapping_result = await db.execute(
                select(models.VariableTagMapping).where(
                    models.VariableTagMapping.variable_id == variable.variable_id,
                    models.VariableTagMapping.context_tag_id == context_tag_id
                )
            )
            mapping = mapping_result.scalars().first()
            
            if mapping:
                var_data["subgroup_tag_id"] = mapping.subgroup_tag_id
                
                # Get tag name
                tag_result = await db.execute(
                    select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
                )
                tag = tag_result.scalars().first()
                if tag:
                    var_data["subgroup_tag_name"] = tag.subgroup_tag_name
        
        response_variables.append(var_data)
    
    return response_variables

@app.get("/api/formula-variables/{variable_id}")
async def get_variable_by_id(variable_id: int, db: AsyncSession = Depends(get_db)):
    # Find the variable
    var_result = await db.execute(
        select(models.FormulaVariable).where(models.FormulaVariable.variable_id == variable_id)
    )
    variable = var_result.scalars().first()
    if not variable:
        raise HTTPException(status_code=404, detail="Variable not found")
    
    # Prepare response data
    response_data = {
        "variable_id": variable.variable_id,
        "variable_name": variable.variable_name,
        "formula_id": variable.formula_id,
        "subgroup_tag_id": variable.subgroup_tag_id,
        "subgroup_tag_name": None
    }
    
    # Get tag name if mapped
    if variable.subgroup_tag_id:
        tag_result = await db.execute(
            select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == variable.subgroup_tag_id)
        )
        tag = tag_result.scalars().first()
        if tag:
            response_data["subgroup_tag_name"] = tag.subgroup_tag_name
    
    return response_data

# Request model for variable mapping
class VariableMappingRequest(BaseModel):
    variable_id: int
    subgroup_tag_id: int  # The tag to map TO
    context_tag_id: int   # The context WHERE the mapping applies

# Create or update variable mapping
@app.put("/api/variable-mappings")
async def create_or_update_variable_mapping(
    mapping: VariableMappingRequest,
    db: AsyncSession = Depends(get_db)
):
    # Check if variable exists
    var_result = await db.execute(
        select(models.FormulaVariable).where(models.FormulaVariable.variable_id == mapping.variable_id)
    )
    variable = var_result.scalars().first()
    if not variable:
        raise HTTPException(status_code=404, detail="Variable not found")
    
    # Check if target tag exists
    tag_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
    )
    tag = tag_result.scalars().first()
    if not tag:
        raise HTTPException(status_code=404, detail="Target tag not found")
    
    # Check if context tag exists
    context_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.context_tag_id)
    )
    context_tag = context_result.scalars().first()
    if not context_tag:
        raise HTTPException(status_code=404, detail="Context tag not found")
    
    # Check for existing mapping
    existing_result = await db.execute(
        select(models.VariableTagMapping).where(
            models.VariableTagMapping.variable_id == mapping.variable_id,
            models.VariableTagMapping.context_tag_id == mapping.context_tag_id
        )
    )
    existing_mapping = existing_result.scalars().first()
    
    if existing_mapping:
        # Update existing mapping
        existing_mapping.subgroup_tag_id = mapping.subgroup_tag_id
        await db.commit()
        await db.refresh(existing_mapping)
        
        return {
            "mapping_id": existing_mapping.mapping_id,
            "variable_id": mapping.variable_id,
            "variable_name": variable.variable_name,
            "subgroup_tag_id": mapping.subgroup_tag_id,
            "subgroup_tag_name": tag.subgroup_tag_name,
            "context_tag_id": mapping.context_tag_id,
            "context_tag_name": context_tag.subgroup_tag_name
        }
    else:
        # Create new mapping
        new_mapping = models.VariableTagMapping(
            variable_id=mapping.variable_id,
            subgroup_tag_id=mapping.subgroup_tag_id,
            context_tag_id=mapping.context_tag_id
        )
        db.add(new_mapping)
        await db.commit()
        await db.refresh(new_mapping)
        
        return {
            "mapping_id": new_mapping.mapping_id,
            "variable_id": mapping.variable_id,
            "variable_name": variable.variable_name,
            "subgroup_tag_id": mapping.subgroup_tag_id,
            "subgroup_tag_name": tag.subgroup_tag_name,
            "context_tag_id": mapping.context_tag_id,
            "context_tag_name": context_tag.subgroup_tag_name
        }

# Get variable mapping for a specific context
@app.get("/api/variable-mappings/{variable_id}")
async def get_variable_mapping(
    variable_id: int,
    context_tag_id: int,
    db: AsyncSession = Depends(get_db)
):
    # Find the mapping
    mapping_result = await db.execute(
        select(models.VariableTagMapping).where(
            models.VariableTagMapping.variable_id == variable_id,
            models.VariableTagMapping.context_tag_id == context_tag_id
        )
    )
    mapping = mapping_result.scalars().first()
    
    if not mapping:
        return {
            "variable_id": variable_id,
            "context_tag_id": context_tag_id,
            "has_mapping": False
        }
    
    # Get variable info
    var_result = await db.execute(
        select(models.FormulaVariable).where(models.FormulaVariable.variable_id == variable_id)
    )
    variable = var_result.scalars().first()
    
    # Get tag info
    tag_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
    )
    tag = tag_result.scalars().first()
    
    return {
        "mapping_id": mapping.mapping_id,
        "variable_id": variable_id,
        "variable_name": variable.variable_name if variable else None,
        "subgroup_tag_id": mapping.subgroup_tag_id,
        "subgroup_tag_name": tag.subgroup_tag_name if tag else None,
        "context_tag_id": mapping.context_tag_id,
        "has_mapping": True
    }

# Get all variable mappings for a context
@app.get("/api/subgroup-tags/{context_tag_id}/variable-mappings")
async def get_context_variable_mappings(
    context_tag_id: int,
    db: AsyncSession = Depends(get_db)
):
    # Check if context tag exists
    context_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == context_tag_id)
    )
    context_tag = context_result.scalars().first()
    if not context_tag:
        raise HTTPException(status_code=404, detail="Context tag not found")
    
    # Find all mappings for this context
    mappings_result = await db.execute(
        select(models.VariableTagMapping).where(
            models.VariableTagMapping.context_tag_id == context_tag_id
        )
    )
    mappings = mappings_result.scalars().all()
    
    result = []
    for mapping in mappings:
        # Get variable info
        var_result = await db.execute(
            select(models.FormulaVariable).where(models.FormulaVariable.variable_id == mapping.variable_id)
        )
        variable = var_result.scalars().first()
        
        # Get tag info
        tag_result = await db.execute(
            select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
        )
        tag = tag_result.scalars().first()
        
        result.append({
            "mapping_id": mapping.mapping_id,
            "variable_id": mapping.variable_id,
            "variable_name": variable.variable_name if variable else None,
            "subgroup_tag_id": mapping.subgroup_tag_id,
            "subgroup_tag_name": tag.subgroup_tag_name if tag else None,
            "context_tag_id": context_tag_id
        })
    
    return result

class TemplateCreate(BaseModel):
    template_name: str
    formula_id: int  # Source formula to copy from

@app.post("/api/templates")
async def create_template(template_data: TemplateCreate, context_tag_id: Optional[int] = None, db: AsyncSession = Depends(get_db)):
    # Verify the source formula exists
    source_formula_result = await db.execute(
        select(models.Formulas).where(models.Formulas.formula_id == template_data.formula_id)
    )
    source_formula = source_formula_result.scalars().first()
    if not source_formula:
        raise HTTPException(status_code=404, detail="Source formula not found")
    
    try:
        # 1. Create a new formula as a copy of the source
        new_formula = models.Formulas(
            formula_name=f"{source_formula.formula_name} (from template {template_data.template_name})",
            formula_desc=source_formula.formula_desc,
            formula_expression=source_formula.formula_expression
        )
        db.add(new_formula)
        await db.flush()
        
        # 2. Create the template linked to the new formula
        new_template = models.Templates(
            template_name=template_data.template_name,
            formula_id=new_formula.formula_id
        )
        db.add(new_template)
        await db.flush()
        
        # 3. Create a special context tag for this template 
        template_context_tag = models.SubgroupTag(
            subgroup_id=None,
            tag_id=None,
            subgroup_tag_name=f"__TEMPLATE_CONTEXT__{new_template.template_id}",
            formula_id=new_formula.formula_id
        )
        db.add(template_context_tag)
        await db.flush()
        
        # 4. Copy all variables from the source formula
        variable_mapping = {}
        source_variables_result = await db.execute(
            select(models.FormulaVariable).where(
                models.FormulaVariable.formula_id == template_data.formula_id
            )
        )
        source_variables = source_variables_result.scalars().all()
        
        for source_var in source_variables:
            new_var = models.FormulaVariable(
                formula_id=new_formula.formula_id,
                variable_name=source_var.variable_name
            )
            db.add(new_var)
            await db.flush()
            variable_mapping[source_var.variable_id] = new_var.variable_id
        
        # 5. Copy variable tag mappings from the specified context
        # Track which variables already have mappings to avoid duplicates
        variables_with_mappings = set()
        
        for source_var_id, new_var_id in variable_mapping.items():
            # Skip if we've already created a mapping for this variable
            if new_var_id in variables_with_mappings:
                continue
                
            mappings_query = select(models.VariableTagMapping).where(
                models.VariableTagMapping.variable_id == source_var_id
            )
            
            # If context_tag_id is provided, only copy mappings from that context
            if context_tag_id is not None:
                mappings_query = mappings_query.where(
                    models.VariableTagMapping.context_tag_id == context_tag_id
                )
                
            mappings_result = await db.execute(mappings_query)
            mappings = mappings_result.scalars().all()
            
            if mappings:
                # Just take the first mapping if multiple exist
                mapping = mappings[0]
                
                # Create new mapping
                new_mapping = models.VariableTagMapping(
                    variable_id=new_var_id,
                    subgroup_tag_id=mapping.subgroup_tag_id,
                    context_tag_id=template_context_tag.subgroup_tag_id
                )
                db.add(new_mapping)
                
                # Mark this variable as having a mapping
                variables_with_mappings.add(new_var_id)
        
        # Commit the transaction
        await db.commit()
        
        return {
            "template_id": new_template.template_id,
            "template_name": new_template.template_name,
            "formula_id": new_formula.formula_id,
            "message": "Template created successfully"
        }
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create template: {str(e)}")
        
class AssignTemplateRequest(BaseModel):
    template_id: int
    subgroup_tag_id: int

# Assign a template's formula to a subgroup tag
@app.post("/api/subgroup-tags/assign-template")
async def assign_template_to_subgroup_tag(
    assignment: AssignTemplateRequest,
    db: AsyncSession = Depends(get_db)
):
    # Verify template exists
    template_result = await db.execute(
        select(models.Templates).where(models.Templates.template_id == assignment.template_id)
    )
    template = template_result.scalars().first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Verify subgroup tag exists
    tag_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == assignment.subgroup_tag_id)
    )
    subgroup_tag = tag_result.scalars().first()
    if not subgroup_tag:
        raise HTTPException(status_code=404, detail="Subgroup tag not found")
    
    try:
        # Get subgroup ID from tag
        subgroup_id = subgroup_tag.subgroup_id
        
        # Update the subgroup tag to use the formula from the template
        subgroup_tag.formula_id = template.formula_id
        await db.flush()
        
        # Create SubgroupTemplate association if it doesn't exist
        if subgroup_id is not None:
            existing_assignment = await db.execute(
                select(models.SubgroupTemplate).where(
                    models.SubgroupTemplate.subgroup_id == subgroup_id,
                    models.SubgroupTemplate.template_id == assignment.template_id
                )
            )
            if not existing_assignment.scalars().first():
                new_assignment = models.SubgroupTemplate(
                    subgroup_id=subgroup_id,
                    template_id=assignment.template_id
                )
                db.add(new_assignment)
                await db.flush()
        
        # Find the template context tag
        context_tag_result = await db.execute(
            select(models.SubgroupTag).where(
                models.SubgroupTag.subgroup_tag_name == f"__TEMPLATE_CONTEXT__{assignment.template_id}"
            )
        )
        template_context_tag = context_tag_result.scalars().first()
        
        if not template_context_tag:
            raise HTTPException(status_code=404, detail="Template context tag not found")
        
        # Find all variables for the template's formula
        variables_result = await db.execute(
            select(models.FormulaVariable).where(
                models.FormulaVariable.formula_id == template.formula_id
            )
        )
        formula_variables = variables_result.scalars().all()
        
        # Delete existing mappings for this context
        for variable in formula_variables:
            await db.execute(
                delete(models.VariableTagMapping).where(
                    models.VariableTagMapping.variable_id == variable.variable_id,
                    models.VariableTagMapping.context_tag_id == assignment.subgroup_tag_id
                )
            )
        
        # Copy mappings from the template's context
        for variable in formula_variables:
            # Check if a new tag was selected for this variable in the new subgroup tag
            selected_mapping_result = await db.execute(
                select(models.VariableTagMapping).where(
                    models.VariableTagMapping.variable_id == variable.variable_id,
                    models.VariableTagMapping.context_tag_id == assignment.subgroup_tag_id
                )
            )
            selected_mapping = selected_mapping_result.scalars().first()
            
            if selected_mapping:
                # Use the selected tag for the new mapping
                new_mapping = models.VariableTagMapping(
                    variable_id=variable.variable_id,
                    subgroup_tag_id=selected_mapping.subgroup_tag_id,
                    context_tag_id=assignment.subgroup_tag_id  # Assign to the new context
                )
            else:
                # Fallback to the template's context mapping
                source_mappings_result = await db.execute(
                    select(models.VariableTagMapping).where(
                        models.VariableTagMapping.variable_id == variable.variable_id,
                        models.VariableTagMapping.context_tag_id == template_context_tag.subgroup_tag_id
                    )
                )
                template_mappings = source_mappings_result.scalars().all()
                
                if template_mappings:
                    mapping = template_mappings[0]
                    new_mapping = models.VariableTagMapping(
                        variable_id=variable.variable_id,
                        subgroup_tag_id=mapping.subgroup_tag_id,
                        context_tag_id=assignment.subgroup_tag_id  # Assign to the new context
                    )
                else:
                    continue  # Skip if no mapping exists
                
            db.add(new_mapping)
            await db.flush()
        
        await db.commit()
        
        return {
            "message": "Template assigned successfully",
            "template_id": assignment.template_id,
            "subgroup_tag_id": assignment.subgroup_tag_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to assign template: {str(e)}")

@app.get("/api/templates")
async def get_all_templates(db: AsyncSession = Depends(get_db)):
    """Get all templates in the system with their variables and mappings"""
    
    result = await db.execute(select(models.Templates))
    templates = result.scalars().all()
    
    template_list = []
    for template in templates:
        # Get formula info
        formula_result = await db.execute(
            select(models.Formulas).where(models.Formulas.formula_id == template.formula_id)
        )
        formula = formula_result.scalars().first()
        
        # Get variables for this formula
        variables_data = []
        if formula:
            variables_result = await db.execute(
                select(models.FormulaVariable).where(models.FormulaVariable.formula_id == formula.formula_id)
            )
            variables = variables_result.scalars().all()
            
            # Get mappings for each variable
            for variable in variables:
                mappings_result = await db.execute(
                    select(models.VariableTagMapping).where(
                        models.VariableTagMapping.variable_id == variable.variable_id
                    )
                )
                mappings = mappings_result.scalars().all()
                
                # Format mappings data
                mappings_data = []
                for mapping in mappings:
                    tag_result = await db.execute(
                        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
                    )
                    tag = tag_result.scalars().first()
                    
                    context_tag_result = await db.execute(
                        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.context_tag_id)
                    )
                    context_tag = context_tag_result.scalars().first()
                    
                    mappings_data.append({
                        "mapping_id": mapping.mapping_id,
                        "subgroup_tag_id": mapping.subgroup_tag_id,
                        "subgroup_tag_name": tag.subgroup_tag_name if tag else None,
                        "context_tag_id": mapping.context_tag_id,
                        "context_tag_name": context_tag.subgroup_tag_name if context_tag else None
                    })
                
                variables_data.append({
                    "variable_id": variable.variable_id,
                    "variable_name": variable.variable_name,
                    "mappings": mappings_data
                })
        
        template_list.append({
            "template_id": template.template_id,
            "template_name": template.template_name,
            "formula_id": template.formula_id,
            "formula_name": formula.formula_name if formula else None,
            "variables": variables_data
        })
    
    return template_list

@app.get("/api/templates/{template_id}")
async def get_template(template_id: int, db: AsyncSession = Depends(get_db)):
    """Get a template with its formula, variables, and tag mappings"""
    
    # Get template
    template_result = await db.execute(
        select(models.Templates).where(models.Templates.template_id == template_id)
    )
    template = template_result.scalars().first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Get formula
    formula_result = await db.execute(
        select(models.Formulas).where(models.Formulas.formula_id == template.formula_id)
    )
    formula = formula_result.scalars().first()
    if not formula:
        raise HTTPException(status_code=404, detail="Template formula not found")
    
    # Get variables
    variables_result = await db.execute(
        select(models.FormulaVariable).where(models.FormulaVariable.formula_id == formula.formula_id)
    )
    variables = variables_result.scalars().all()
    
    # Build variables data with their mappings
    variables_data = []
    for variable in variables:
        # Get mappings for this variable
        mappings_result = await db.execute(
            select(models.VariableTagMapping).where(
                models.VariableTagMapping.variable_id == variable.variable_id
            )
        )
        mappings = mappings_result.scalars().all()
        
        # Format mappings data
        mappings_data = []
        for mapping in mappings:
            tag_result = await db.execute(
                select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.subgroup_tag_id)
            )
            tag = tag_result.scalars().first()
            
            context_tag_result = await db.execute(
                select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == mapping.context_tag_id)
            )
            context_tag = context_tag_result.scalars().first()
            
            mappings_data.append({
                "mapping_id": mapping.mapping_id,
                "subgroup_tag_id": mapping.subgroup_tag_id,
                "subgroup_tag_name": tag.subgroup_tag_name if tag else None,
                "context_tag_id": mapping.context_tag_id,
                "context_tag_name": context_tag.subgroup_tag_name if context_tag else None
            })
        
        variables_data.append({
            "variable_id": variable.variable_id,
            "variable_name": variable.variable_name,
            "mappings": mappings_data
        })
    
    # Return template with all associated data
    return {
        "template_id": template.template_id,
        "template_name": template.template_name,
        "formula": {
            "formula_id": formula.formula_id,
            "formula_name": formula.formula_name,
            "formula_desc": formula.formula_desc,
            "formula_expression": formula.formula_expression
        },
        "variables": variables_data
    }  

@app.post("/api/subgroup-tags/assign-template")
async def assign_template_to_subgroup_tag(
    assignment: AssignTemplateRequest,
    db: AsyncSession = Depends(get_db)
):
    # Verify template exists
    template_result = await db.execute(
        select(models.Templates).where(models.Templates.template_id == assignment.template_id)
    )
    template = template_result.scalars().first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Verify subgroup tag exists
    tag_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == assignment.subgroup_tag_id)
    )
    subgroup_tag = tag_result.scalars().first()
    if not subgroup_tag:
        raise HTTPException(status_code=404, detail="Subgroup tag not found")
    
    try:
        # Get subgroup ID from tag
        subgroup_id = subgroup_tag.subgroup_id
        
        # Update the subgroup tag to use the formula from the template
        subgroup_tag.formula_id = template.formula_id
        await db.flush()
        
        # Create SubgroupTemplate association if it doesn't exist
        if subgroup_id is not None:
            existing_assignment = await db.execute(
                select(models.SubgroupTemplate).where(
                    models.SubgroupTemplate.subgroup_id == subgroup_id,
                    models.SubgroupTemplate.template_id == assignment.template_id
                )
            )
            if not existing_assignment.scalars().first():
                new_assignment = models.SubgroupTemplate(
                    subgroup_id=subgroup_id,
                    template_id=assignment.template_id
                )
                db.add(new_assignment)
                await db.flush()
        
        # Find the template context tag
        context_tag_result = await db.execute(
            select(models.SubgroupTag).where(
                models.SubgroupTag.subgroup_tag_name == f"__TEMPLATE_CONTEXT__{assignment.template_id}"
            )
        )
        template_context_tag = context_tag_result.scalars().first()
        
        if not template_context_tag:
            raise HTTPException(status_code=404, detail="Template context tag not found")
        
        # Find all variables for the template's formula
        variables_result = await db.execute(
            select(models.FormulaVariable).where(
                models.FormulaVariable.formula_id == template.formula_id
            )
        )
        formula_variables = variables_result.scalars().all()
        
        # Delete existing mappings for this context
        for variable in formula_variables:
            await db.execute(
                delete(models.VariableTagMapping).where(
                    models.VariableTagMapping.variable_id == variable.variable_id,
                    models.VariableTagMapping.context_tag_id == assignment.subgroup_tag_id
                )
            )
        
        # Track variables that already have mappings
        variables_with_mappings = set()
        
        # Copy mappings from the template's context
        for variable in formula_variables:
            # Skip if we've already created a mapping for this variable
            if variable.variable_id in variables_with_mappings:
                continue
            
            source_mappings_result = await db.execute(
                select(models.VariableTagMapping).where(
                    models.VariableTagMapping.variable_id == variable.variable_id,
                    models.VariableTagMapping.context_tag_id == template_context_tag.subgroup_tag_id
                )
            )
            template_mappings = source_mappings_result.scalars().all()
            
            if template_mappings:
                # Just take the first mapping to avoid duplicates
                mapping = template_mappings[0]
                
                # Create new mapping using this tag
                new_mapping = models.VariableTagMapping(
                    variable_id=variable.variable_id,
                    subgroup_tag_id=mapping.subgroup_tag_id,
                    context_tag_id=assignment.subgroup_tag_id  # Assign to the new context
                )
                db.add(new_mapping)
                await db.flush()
                
                # Mark this variable as having a mapping
                variables_with_mappings.add(variable.variable_id)
        
        await db.commit()
        
        return {
            "message": "Template assigned successfully",
            "template_id": assignment.template_id,
            "subgroup_tag_id": assignment.subgroup_tag_id
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to assign template: {str(e)}")
    
@app.delete("/api/templates/{template_id}")
async def delete_template(template_id: int, db: AsyncSession = Depends(get_db)):
    # Verify template exists
    template_result = await db.execute(
        select(models.Templates).where(models.Templates.template_id == template_id)
    )
    template = template_result.scalars().first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    try:
        # Delete the template's context tag
        context_tag_result = await db.execute(
            select(models.SubgroupTag).where(
                models.SubgroupTag.subgroup_tag_name == f"__TEMPLATE_CONTEXT__{template_id}"
            )
        )
        context_tag = context_tag_result.scalars().first()
        if context_tag:
            await db.execute(
                delete(models.SubgroupTag).where(
                    models.SubgroupTag.subgroup_tag_id == context_tag.subgroup_tag_id
                )
            )
        
        # Delete the template itself
        await db.execute(
            delete(models.Templates).where(models.Templates.template_id == template_id)
        )
        
        await db.commit()
        return {"message": "Template deleted successfully"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete template: {str(e)}")

@app.delete("/api/subgroup-tags/{subgroup_tag_id}")
async def delete_subgroup_tag(subgroup_tag_id: int, db: AsyncSession = Depends(get_db)):
    # Verify subgroup tag exists
    tag_result = await db.execute(
        select(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == subgroup_tag_id)
    )
    subgroup_tag = tag_result.scalars().first()
    if not subgroup_tag:
        raise HTTPException(status_code=404, detail="Subgroup tag not found")
    
    try:
        # Delete all mappings associated with this subgroup tag
        await db.execute(
            delete(models.VariableTagMapping).where(
                models.VariableTagMapping.context_tag_id == subgroup_tag_id
            )
        )
        
        # Delete the subgroup tag itself
        await db.execute(
            delete(models.SubgroupTag).where(models.SubgroupTag.subgroup_tag_id == subgroup_tag_id)
        )
        
        await db.commit()
        return {"message": "Subgroup tag deleted successfully"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete subgroup tag: {str(e)}")
    
@app.delete("/api/variable-mappings/{mapping_id}")
async def delete_variable_mapping(mapping_id: int, db: AsyncSession = Depends(get_db)):
    # Verify mapping exists
    mapping_result = await db.execute(
        select(models.VariableTagMapping).where(models.VariableTagMapping.mapping_id == mapping_id)
    )
    mapping = mapping_result.scalars().first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Variable mapping not found")
    
    try:
        # Delete the mapping
        await db.execute(
            delete(models.VariableTagMapping).where(models.VariableTagMapping.mapping_id == mapping_id)
        )
        
        await db.commit()
        return {"message": "Variable mapping deleted successfully"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete variable mapping: {str(e)}")

@app.delete("/api/formulas/{formula_id}")
async def delete_formula(formula_id: int, db: AsyncSession = Depends(get_db)):
    # Verify formula exists
    formula_result = await db.execute(
        select(models.Formulas).where(models.Formulas.formula_id == formula_id)
    )
    formula = formula_result.scalars().first()
    if not formula:
        raise HTTPException(status_code=404, detail="Formula not found")
    
    try:
        # Delete all variables associated with this formula
        await db.execute(
            delete(models.FormulaVariable).where(
                models.FormulaVariable.formula_id == formula_id
            )
        )
        
        # Delete the formula itself
        await db.execute(
            delete(models.Formulas).where(models.Formulas.formula_id == formula_id)
        )
        
        await db.commit()
        return {"message": "Formula deleted successfully"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete formula: {str(e)}")

@app.delete("/api/subgroups/{subgroup_id}")
async def delete_subgroup(subgroup_id: int, db: AsyncSession = Depends(get_db)):
    # Verify subgroup exists
    subgroup_result = await db.execute(
        select(models.Subgroups).where(models.Subgroups.subgroup_id == subgroup_id)
    )
    subgroup = subgroup_result.scalars().first()
    if not subgroup:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    try:
        # Delete all tags associated with this subgroup
        await db.execute(
            delete(models.SubgroupTag).where(models.SubgroupTag.subgroup_id == subgroup_id)
        )
        
        # Delete the subgroup itself
        await db.execute(
            delete(models.Subgroups).where(models.Subgroups.subgroup_id == subgroup_id)
        )
        
        await db.commit()
        return {"message": "Subgroup deleted successfully"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete subgroup: {str(e)}")